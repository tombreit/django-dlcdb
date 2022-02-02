import datetime

from copy import copy
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Color, Alignment, colors, PatternFill
from openpyxl.utils import get_column_letter

from django.core.files.base import ContentFile
from django.utils.text import slugify

from ..settings import EXPOSED_FIELDS


def get_record_data_row(record, event):
    """
    Build the data part (rows) for a given record and an event.
    """
    row = []
    for item in EXPOSED_FIELDS:
        if event in item.get("used_for"):
            obj = record
            for model_name in item.get("model"):
                obj = getattr(obj, model_name)
            field = getattr(obj, item.get("field"))
            if isinstance(field, datetime.date):
                field = '{:%Y-%m-%d}'.format(field)
            row.append(field)

    return row


def get_records_as_text(records=None, title=None, event=None):
    """
    Build a string representation of affected records. Could be used as an email body.
    """
    result = []
    for record in records:
        result.append(get_record_data_row(record, event))

    return result


def get_records_as_spreadsheet(records=None, title=None, event=None):
    """
    Return records wrapped in a spreadsheet file (xlsx).
    """

    with NamedTemporaryFile(suffix='xlsx') as tmp:

        # Creating xlsx via openpyxl
        workbook = Workbook()

        _base_font = Font(name="DejaVu Sans Mono")  # Nimbus Mono

        title_style = NamedStyle(name="title_style")
        title_style.font = copy(_base_font)
        title_style.font = Font(bold=True, color='FFFFFFFF')
        title_style.alignment = Alignment(horizontal='left', vertical='center')
        title_style.fill = PatternFill(fill_type='solid', bgColor='000000')

        header_style = NamedStyle(name="header_style")
        header_style.font = copy(_base_font)
        header_style.font = Font(bold=True)
        header_style.alignment = Alignment(horizontal='right', vertical='top')

        body_style = NamedStyle(name="body_style")
        body_style.font = copy(_base_font)
        body_style.alignment = Alignment(horizontal='left', vertical='top')

        ws = workbook.active
        ws.title = slugify(title) if title else ''

        # row 1: title
        # _title = title if title else ''
        # ws.append([_title])
        # # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
        # ws.merge_cells('A1:X1')
        # title_row = ws[1]
        # for cell in title_row:
        #     cell.style = title_style

        # row 1: -blank-
        ws.append([])
        
        # row 2: -blank-
        ws.append([])
        
        # row 3: column names
        # ws.append(EXPOSED_FIELDS_RECORD + EXPOSED_FIELDS_DEVICE)
        col_headers_row = []
        for item in EXPOSED_FIELDS:
            if event in item.get("used_for"):
                header_str = '{model_name}-{field}'.format(
                    model_name=''.join(item.get("model")) if item.get("model") else 'Record',
                    field=item.get("field"),
                )
                col_headers_row.append(header_str)
        ws.append(col_headers_row)

        column_header_row = ws[3]
        for cell in column_header_row:
            cell.style = header_style

        # row 4: -blank-
        ws.append([])

        # row 5+: data
        for record in records:
            ws.append(get_record_data_row(record, event))

        # for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        #     print('Row number:', str(row[0].row))
        #     row[0].style = body

        # for row_cells in ws.iter_rows(min_row=ws._current_row, max_row=ws._current_row):
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.style = body_style


        # Optimized column widths
        # see: https://stackoverflow.com/questions/13197574/openpyxl-adjust-column-width-size
        column_widths = []
        for row in ws.iter_rows(min_row=3):
            for i, cell in enumerate(row):
                if cell.value:
                    try:
                        column_widths[i] = max(column_widths[i], len(cell.value))
                    except IndexError:
                        column_widths.append(len(cell.value))

        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width

        workbook.save(tmp.name)
        content_file_obj = ContentFile(tmp.read())
        return content_file_obj
